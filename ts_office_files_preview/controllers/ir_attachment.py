from odoo import http
from odoo.http import request
import base64
import csv
import io
import json
import openpyxl
import tempfile
import subprocess
import os


class AttachmentPreviewController(http.Controller):

    @http.route('/csv/preview/<int:attachment_id>', auth='user', type='http')
    def preview_attachment(self, attachment_id):

        attachment = request.env['ir.attachment'].sudo().browse(attachment_id)

        if not attachment.exists() or not attachment.datas:
            return self._json_response({"sheets": []})

        file_data = base64.b64decode(attachment.datas)
        filename = attachment.name.lower()

        try:
            file_input = io.BytesIO(file_data)

            # XLSX Preview (Multiple Sheets)
            if filename.endswith(".xlsx"):

                wb = openpyxl.load_workbook(
                    file_input,
                    read_only=True,
                    data_only=True
                )

                sheets_data = []
                MAX_ROWS = 500  # prevent huge files freezing UI

                for sheet_name in wb.sheetnames:

                    sheet = wb[sheet_name]
                    rows = []

                    for i, row in enumerate(sheet.iter_rows(values_only=True)):

                        if i >= MAX_ROWS:
                            break

                        rows.append([
                            str(cell) if cell is not None else ""
                            for cell in row
                        ])

                    sheets_data.append({
                        "name": sheet_name,
                        "rows": rows
                    })

                return self._json_response({
                    "sheets": sheets_data
                })

            # CSV Preview
            elif filename.endswith(".csv"):

                file_content = file_data.decode("utf-8", errors="ignore")
                csv_file = io.StringIO(file_content)

                try:
                    dialect = csv.Sniffer().sniff(file_content[:1024]) if file_content else csv.excel
                    reader = csv.reader(csv_file, dialect)
                except Exception:
                    csv_file.seek(0)
                    reader = csv.reader(csv_file)

                rows = list(reader)

                return self._json_response({
                    "sheets": [
                        {
                            "name": attachment.name,
                            "rows": rows
                        }
                    ]
                })

        except Exception as e:
            return self._json_response({"error": str(e)})

    # @http.route('/ppt/preview/<int:attachment_id>', auth='user', type='http')
    # def ppt_preview(self, attachment_id):
    #
    #     attachment = request.env['ir.attachment'].sudo().browse(attachment_id)
    #
    #
    #     print(attachment)
    #
    #     if not attachment.exists() or not attachment.datas:
    #         return request.not_found()
    #
    #     try:
    #
    #         file_data = base64.b64decode(attachment.datas)
    #
    #         with tempfile.TemporaryDirectory() as tmpdir:
    #
    #             ppt_path = os.path.join(tmpdir, attachment.name)
    #             pdf_path = ppt_path.replace(".pptx", ".pdf")
    #
    #             # Save PPT file
    #             with open(ppt_path, "wb") as f:
    #                 f.write(file_data)
    #
    #             # Convert PPT → PDF
    #             subprocess.run([
    #                 r"C:\Program Files\LibreOffice\program\soffice.exe",
    #                 "--headless",
    #                 "--convert-to", "pdf",
    #                 ppt_path,
    #                 "--outdir", tmpdir
    #             ], check=True)
    #
    #             # Return PDF
    #             with open(pdf_path, "rb") as pdf:
    #                 content = pdf.read()
    #
    #             return request.make_response(
    #                 content,
    #                 headers=[
    #                     ("Content-Type", "application/pdf"),
    #                     ("Content-Length", len(content))
    #                 ]
    #             )
    #
    #     except Exception as e:
    #         return request.make_response(
    #             str(e),
    #             headers=[("Content-Type", "text/plain")]
    #         )

    def _json_response(self, data):
        return request.make_response(
            json.dumps(data),
            headers=[("Content-Type", "application/json")]
        )
